import enum
from .models import Metadata, ScienceKeyword, Label
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def get_row_data_columns(file):
    wb = load_workbook(file)
    ws = wb[wb.sheetnames[0]]

    auto_delete_null_rows(ws)

    rows = ws.rows
    head = next(rows)

    titles = []
    for cell in head:
        if cell.value is not None:
            # print(cell, cell.value)
            titles.append(cell.value.lower())

    return titles


def generate_science_keywords(file, keywords, least_time=1, excludes=None):
    titles = get_row_data_columns(file)
    # print(titles)
    if excludes:
        titles = filter(lambda x: x not in excludes, titles)

    keyword_time_dic = {}

    labels = Label.objects.filter(name__in=titles)

    for label in labels:
        for keyword in label.keywords.all():
            if keyword in keyword_time_dic:
                keyword_time_dic[keyword] += 1
            else:
                keyword_time_dic[keyword] = 1

    keyword_time_dic = dict(
        filter(lambda x: x[1] >= least_time, keyword_time_dic.items())
    )
    keyword_list = get_sorted_list(keyword_time_dic, True)
    print(keyword_list)

    for keyword, time in keyword_list:
        if isinstance(keywords, list):
            keywords.append(keyword)
        else:
            keywords.add(keyword)


def get_sorted_list(d, reverse=False):
    return sorted(d.items(), key=lambda x: x[1], reverse=reverse)


def import_science_keywords(file):
    wb = load_workbook(file)
    ws = wb[wb.sheetnames[1]]
    rows = ws.rows

    # print(get_valid_max_row(ws))
    # exluce the head
    head = next(rows)

    # extract the keyword
    for id, topic, term, v1, v2, v3 in rows:
        if topic.value is not None:
            keyword = ScienceKeyword.objects.get_or_create(
                topic=topic.value,
                term=term.value,
                variable1=v1.value,
                variable2=v2.value,
                variable3=v3.value,
            )
        else:
            break


def import_science_lables(file):
    wb = load_workbook(file)
    ws = wb[wb.sheetnames[0]]
    rows = ws.rows
    head = next(rows)
    print(head)
    for topic, term, v1, v2, v3, lables in rows:
        if topic.value is not None:
            fullname = get_science_full_name(topic, term, v1, v2, v3)
            try:
                keyword = ScienceKeyword.objects.get(full_name=fullname)
                li = lables.value.split(",")
                for label in li:
                    obj, _ = Label.objects.get_or_create(name=label.lower())
                    obj.keywords.add(keyword)
                    obj.save()
                    print(obj)
            except:
                print("not exist " + fullname)


def get_science_full_name(topic, term, variable1, variable2, variable3):
    str = f"{topic.value}"
    if term.value:
        str += f" > {term.value}"
    if variable1.value:
        str += f" > {variable1.value}"
    if variable2.value:
        str += f"> {variable2.value}"
    if variable3.value:
        str += f" > {variable3.value}"
    return str


##################################################################
def get_valid_max_row(ws):
    for max_row, row in enumerate(ws, 1):
        if all(c.value is None for c in row):
            return max_row


def auto_delete_null_rows(ws, max_col=50):
    empty_rows = []
    for idx, row in enumerate(ws.iter_rows(max_col=max_col), start=1):
        empty = not any((cell.value for cell in row))
        if empty:
            empty_rows.append(idx)
    for row_idx in reversed(empty_rows):
        ws.delete_rows(row_idx, 1)
