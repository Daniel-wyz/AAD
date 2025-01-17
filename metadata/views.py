from curses import meta
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, JsonResponse
from urllib3 import HTTPResponse
from .forms import MetadataForm, MetadataChartForm
from .models import Metadata, RawData, ScienceKeyword
from .utils import (
    import_science_keywords,
    auto_delete_null_rows,
    generate_science_keywords,
    import_science_lables,
    get_row_data_columns,
)
from AAD.utils import is_ajax, generate_code, get_chart
from profiles.models import Profile

from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import ListView, DetailView, TemplateView

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import random
import pandas as pd

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# for pdf
from django.conf import settings
from django.http import HttpResponse
from django.template.loader import get_template
from xhtml2pdf import pisa

import json
from django.db.models import Q
from django.db.models import F


class MetadataListView(LoginRequiredMixin, ListView):
    model = Metadata
    template_name = "metadata/home.html"


@login_required
def metadata_detail_view(request, metadata_id):
    object = Metadata.objects.get(metadata_id=metadata_id)
    return render(request, "metadata/detail.html", {"object": object})


@login_required
def render_pdf_view(request, metadata_id):
    template_path = "metadata/pdf.html"
    object = get_object_or_404(Metadata, metadata_id=metadata_id)
    context = {"object": object}

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'filename="report.pdf"'

    template = get_template(template_path)
    html = template.render(context)

    pisa_status = pisa.CreatePDF(html, dest=response)

    if pisa_status.err:
        return HttpResponse("We had some errors <pre>" + html + "</pre>")
    return response


@login_required
def download_xml(request, metadata_id):
    template_path = "metadata/xml.xml"
    object = get_object_or_404(Metadata, metadata_id=metadata_id)
    template = get_template(template_path)
    xml = template.render({"object": object})
    response = HttpResponse(xml, content_type="application/xml")
    response["Content-Disposition"] = f"attachment;filename={metadata_id}.xml"
    return response


@login_required
def upload_documents(request):
    if request.method == "POST":
        result = {}
        ## 1. create a blank metadata
        author = Profile.objects.get(user=request.user)
        metadata = Metadata.objects.create(
            metadata_id=generate_code(),
            author=author,
        )

        for file in request.FILES.values():
            # import_science_lables(file)
            ## 2.assign rawdata for metadata
            obj, created = RawData.objects.get_or_create(file_name=file.name)
            if created:
                obj.file_obj = file
                obj.save()
            metadata.metadata_name = file.name
            metadata.rawdata.add(obj)

            ## 3.analyze and autofill the suggested science keywords
            generate_science_keywords(file, metadata.science_keywords)

        metadata.save()

        result["success"] = True
        result["metadata_id"] = metadata.metadata_id

        return JsonResponse(result)
    return render(request, "metadata/upload.html", {})


@login_required
def create_metadata(request, metadata_id):
    form = None
    err_msg = None
    success = None
    columns = None
    metadata = Metadata.objects.get(metadata_id=metadata_id)
    if metadata:
        ## not the author
        if metadata.author.user != request.user:
            err_msg = "Only the author can edit metadata"
        else:
            form = MetadataForm(request.POST or None, instance=metadata)

            form.fields["rawdata"].queryset = metadata.rawdata
            form.initial["rawdata"] = metadata.rawdata.all()
            columns = get_row_data_columns(metadata.rawdata.first().file_obj)

            # form.fields[
            #     "science_keywords"
            # ].queryset = metadata.science_keywords.order_by(
            #     "topic", "term", "variable1", "variable2", "variable3"
            # )
            # form.initial["science_keywords"] = metadata.science_keywords.all()

            if request.method == "POST":
                print(request.POST)

                if form.is_valid():
                    keywords = request.POST.getlist("science_keywords")
                    # keywords = form.cleaned_data.get("science_keywords")
                    if len(keywords) > 0:
                        metadata.science_keywords.set(keywords)
                        metadata.save()
                        success = True
                    else:
                        err_msg = "must select at least one keyword"
                else:
                    err_msg = form.errors
    else:
        err_msg = "no metadata can be founded"

    context = {
        "form": form,
        "metadata": metadata,
        "columns": columns,
        "success": success,
        "err_msg": err_msg,
    }
    return render(request, "metadata/create.html", context)


def search_science_keywords(request):
    result = {}
    if request.method == "GET":
        start = int(request.GET.get("iDisplayStart", "0"))
        length = int(request.GET.get("iDisplayLength", "30"))
        search = request.GET.get("search", "")

        if search:
            #  orgs = user_list.objects.filter(Q(topic__icontains=search) | Q ( term__icontains=search)).values('id').annotate(text=F('full_name'))
            # 截取查询结果对象，以start开始截取start+length位
            orgs = (
                ScienceKeyword.objects.filter(full_name__icontains=search)
                .values("id")
                .annotate(text=F("full_name"))
            )
        else:
            orgs = (
                ScienceKeyword.objects.all().values("id").annotate(text=F("full_name"))
            )
        ret = list(orgs)
        result = json.dumps(ret)
    return HttpResponse(result)


def filter_science_keywords(request):
    result = {}
    if request.method == "GET":
        print(request.GET)
        metadata_id = request.GET.get("metadata_id")
        hitcount = int(request.GET.get("hitcount"))
        excludes = request.GET.getlist("excludes")

        keywords = []
        metadata = Metadata.objects.get(metadata_id=metadata_id)
        rawdata = metadata.rawdata.first().file_obj

        generate_science_keywords(rawdata, keywords, hitcount, excludes)
        keywords = map(lambda x: str(x.id) + "-" + x.full_name, keywords)

        result = json.dumps(list(keywords))

    return HttpResponse(result)
